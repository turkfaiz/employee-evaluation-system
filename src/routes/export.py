from flask import Blueprint, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from src.models.user import db
from src.models.employee import Employee
from src.models.evaluation import MonthlyEvaluation, EvaluationScore
from src.models.department import Department, EvaluationCriteria
from datetime import datetime
import os
import tempfile

export_bp = Blueprint('export', __name__)

@export_bp.route('/export/all-evaluations', methods=['GET'])
def export_all_evaluations():
    """تصدير جميع تقييمات الموظفين إلى ملف Excel"""
    try:
        # إنشاء ملف Excel جديد
        wb = Workbook()
        
        # حذف الورقة الافتراضية
        wb.remove(wb.active)
        
        # الحصول على جميع الموظفين
        employees = Employee.query.all()
        
        if not employees:
            return jsonify({'error': 'لا توجد بيانات موظفين للتصدير'}), 400
        
        # إنشاء ورقة عمل لكل موظف
        for employee in employees:
            create_employee_sheet(wb, employee)
        
        # إنشاء ورقة ملخص عام
        create_summary_sheet(wb, employees)
        
        # حفظ الملف في مجلد مؤقت
        temp_dir = tempfile.mkdtemp()
        filename = f"تقييمات_الموظفين_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
        filepath = os.path.join(temp_dir, filename)
        
        wb.save(filepath)
        
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@export_bp.route('/export/employee/<int:employee_id>', methods=['GET'])
def export_employee_evaluations(employee_id):
    """تصدير تقييمات موظف محدد إلى ملف Excel"""
    try:
        employee = Employee.query.get_or_404(employee_id)
        
        # إنشاء ملف Excel جديد
        wb = Workbook()
        wb.remove(wb.active)
        
        # إنشاء ورقة عمل للموظف
        create_employee_sheet(wb, employee)
        
        # حفظ الملف في مجلد مؤقت
        temp_dir = tempfile.mkdtemp()
        filename = f"تقييم_{employee.full_name}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
        filepath = os.path.join(temp_dir, filename)
        
        wb.save(filepath)
        
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def create_employee_sheet(wb, employee):
    """إنشاء ورقة عمل لموظف محدد"""
    # إنشاء ورقة جديدة
    ws = wb.create_sheet(title=employee.full_name[:30])  # تحديد طول الاسم
    
    # تنسيق الخلايا
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # معلومات الموظف
    ws['A1'] = 'معلومات الموظف'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = center_alignment
    ws.merge_cells('A1:D1')
    
    ws['A2'] = 'الاسم الكامل'
    ws['B2'] = employee.full_name
    ws['A3'] = 'الرقم الوظيفي'
    ws['B3'] = employee.employee_number
    ws['A4'] = 'المسمى الوظيفي'
    ws['B4'] = employee.job_title
    ws['A5'] = 'الإدارة'
    ws['B5'] = employee.department.name
    
    # الحصول على معايير التقييم للإدارة
    criteria = EvaluationCriteria.query.filter_by(department_id=employee.department_id).all()
    
    # الحصول على تقييمات الموظف
    evaluations = MonthlyEvaluation.query.filter_by(employee_id=employee.id).order_by(
        MonthlyEvaluation.evaluation_year, MonthlyEvaluation.evaluation_month
    ).all()
    
    if not evaluations:
        ws['A7'] = 'لا توجد تقييمات لهذا الموظف'
        return
    
    # رأس جدول التقييمات
    row = 7
    ws[f'A{row}'] = 'التقييمات الشهرية'
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].alignment = center_alignment
    
    # تحديد عدد الأعمدة (الشهر + السنة + المعايير + المتوسط)
    total_cols = 3 + len(criteria)
    ws.merge_cells(f'A{row}:{get_column_letter(total_cols)}{row}')
    
    # رؤوس الأعمدة
    row += 1
    ws[f'A{row}'] = 'الشهر'
    ws[f'B{row}'] = 'السنة'
    
    col = 3
    for criterion in criteria:
        ws[f'{get_column_letter(col)}{row}'] = criterion.criteria_name
        col += 1
    
    ws[f'{get_column_letter(col)}{row}'] = 'المتوسط'
    
    # تطبيق التنسيق على رؤوس الأعمدة
    for c in range(1, total_cols + 1):
        cell = ws[f'{get_column_letter(c)}{row}']
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # بيانات التقييمات
    months_ar = [
        '', 'يناير', 'فبراير', 'مارس', 'أبريل', 'مايو', 'يونيو',
        'يوليو', 'أغسطس', 'سبتمبر', 'أكتوبر', 'نوفمبر', 'ديسمبر'
    ]
    
    for evaluation in evaluations:
        row += 1
        ws[f'A{row}'] = months_ar[evaluation.evaluation_month]
        ws[f'B{row}'] = evaluation.evaluation_year
        
        # الدرجات لكل معيار
        scores_dict = {score.criteria_id: score.score for score in evaluation.scores}
        col = 3
        total_score = 0
        
        for criterion in criteria:
            score = scores_dict.get(criterion.id, 0)
            ws[f'{get_column_letter(col)}{row}'] = score
            total_score += score
            col += 1
        
        # المتوسط
        average = total_score / len(criteria) if criteria else 0
        ws[f'{get_column_letter(col)}{row}'] = round(average, 2)
        
        # تطبيق التنسيق على الصف
        for c in range(1, total_cols + 1):
            cell = ws[f'{get_column_letter(c)}{row}']
            cell.alignment = center_alignment
            cell.border = border
    
    # تعديل عرض الأعمدة
    for col in range(1, total_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

def create_summary_sheet(wb, employees):
    """إنشاء ورقة ملخص عام لجميع الموظفين"""
    ws = wb.create_sheet(title="الملخص العام", index=0)
    
    # تنسيق الخلايا
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # العنوان الرئيسي
    ws['A1'] = 'ملخص تقييمات جميع الموظفين'
    ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws['A1'].fill = header_fill
    ws['A1'].alignment = center_alignment
    ws.merge_cells('A1:F1')
    
    # رؤوس الأعمدة
    headers = ['الاسم الكامل', 'الرقم الوظيفي', 'المسمى الوظيفي', 'الإدارة', 'عدد التقييمات', 'المتوسط العام']
    for col, header in enumerate(headers, 1):
        cell = ws[f'{get_column_letter(col)}3']
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # بيانات الموظفين
    row = 4
    for employee in employees:
        evaluations = MonthlyEvaluation.query.filter_by(employee_id=employee.id).all()
        
        # حساب المتوسط العام
        total_average = 0
        evaluation_count = len(evaluations)
        
        if evaluation_count > 0:
            for evaluation in evaluations:
                total_average += evaluation.get_average_score()
            total_average = total_average / evaluation_count
        
        # إدخال البيانات
        ws[f'A{row}'] = employee.full_name
        ws[f'B{row}'] = employee.employee_number
        ws[f'C{row}'] = employee.job_title
        ws[f'D{row}'] = employee.department.name
        ws[f'E{row}'] = evaluation_count
        ws[f'F{row}'] = round(total_average, 2) if evaluation_count > 0 else 'لا توجد تقييمات'
        
        # تطبيق التنسيق
        for col in range(1, 7):
            cell = ws[f'{get_column_letter(col)}{row}']
            cell.alignment = center_alignment
            cell.border = border
        
        row += 1
    
    # تعديل عرض الأعمدة
    column_widths = [20, 15, 20, 15, 15, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

